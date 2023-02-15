import os
import numpy as np
import cv2
import torch
from scipy.optimize import linear_sum_assignment
import xlsxwriter
import matplotlib.pyplot as plt
from numpy.linalg import norm
import time
from multiprocessing import Pool
import _thread
from multiprocessing.pool import ThreadPool
from multiprocessing import Process
import sys
from pympler import muppy, summary
#from write_excel import create_excel
from openpyxl import *
from scipy.signal import savgol_filter
import itertools
import pandas as pd
import scipy
from natsort import natsorted

def compute_feature_matching(support_images_all, des, batch_number, batch_images,total_frame,RoI,target,target_all, patch_size=32):

    des = des[RoI]
    f0 = des[0] # empty or full ?
    f1 = des[1]
    i=-1
    
    for img in batch_images[RoI]:
        i+=1
        img=img.numpy().astype('uint8')
        f = np.array([])

        for i in range(3):
            h = cv2.calcHist([img], [i], None, [256], [0,256])
            f = np.append(f, h) 

        d0 = ((f - f0)**2).sum()
        d1 = ((f - f1)**2).sum()

        if d0 < d1:
            target[RoI].append(1)
        else:
            target[RoI].append(0)
       
    return (target)

def memory_stat():
    all_objects = muppy.get_objects()
    sum1 = summary.summarize(all_objects)
    summary.print_(sum1)
    dataframes = [ao for ao in all_objects if isinstance(ao, pd.DataFrame)]
    for d in dataframes:
        print (d.columns.values)
        print (len(d))
    
def write_excel(target,step_frame):
       
    wb=load_workbook("D:/OtherCodes/BoxTracking/Station_Classifier/box_realtime_schedule.xlsx")
    Excel_file=wb["Sheet1"]
    row=3
    column=8
    start=[]
    finish=[]
    duration=[]

    def track(target):
        State_Change_List = []
        for State_Change_Index, item in enumerate(target):
            if State_Change_Index != 0:
                if target[State_Change_Index] != target[State_Change_Index-1]:
                    State_Change_List.append(State_Change_Index)

        return (State_Change_List)
        
    State_Change_List=track(target)

    for i in range(len(State_Change_List)):
        cell_start=Excel_file.cell(row,column)        
        if target[State_Change_List[i]]==1:
            start.append(State_Change_List[i])
        else:
            finish.append(State_Change_List[i])        

    for row in range(len(start)):
        cell_start=Excel_file.cell(row+3,2)  
        cell_start.value=start[row]*step_frame
        
    for row in range(len(finish)):    
        cell_finish=Excel_file.cell(row+3,3)
        cell_finish.value=finish[row]*step_frame
        
    wb.save("D:/OtherCodes/BoxTracking/Station_Classifier/box_realtime_schedule_updated.xlsx")
    
def write_video(input_dir, output_dir, target):
    i=0
    print(target)
    box=[]
    box.append([850,320,1300,1220])
    box.append([200,170,720,1220])
    
    fourcc = cv2.VideoWriter_fourcc('M','J','P','G')
    video=input_dir
    cap = cv2.VideoCapture(video)
    ret, frame = cap.read()
    
    fps = cap.get(cv2.CAP_PROP_FPS)
    write_total_frames=cap.get(cv2.CAP_PROP_FRAME_COUNT)
    m,n,_=frame.shape
    video_writer = cv2.VideoWriter(output_dir, fourcc, fps, (n,m ))

    if draw_cost_bool==True:
        del cost[0][-1]
        del cost[0][-1]
        
        plt.plot(cost[0])
        plt.show()
        plt.plot(cost[1])
        plt.show()
   
    counter=0
    print(write_total_frames)
    while(counter<write_total_frames):
        #print(counter)
        status=[]
        counter_step=counter//step_frame
        #print(counter_step)
        ret, frame = cap.read()
        if counter%step_frame==0:
            if ret==True:
                for RoI in range(RoI_Count):
                    RoI=0
                    if target[RoI][counter_step]==0:
                        status.append("{} is empty".format(Stations[RoI]))
                        color=(0,255,0)
                    elif target[RoI][counter_step]==1:
                        status.append("{} is occupied".format(Stations[RoI]))          
                        color=(0, 0, 200)

                # The big rectangle 
                    frame = cv2.rectangle(frame, (box[RoI][0],box[RoI][1]), (box[RoI][2],box[RoI][3]), color, 2)      
                    cv2.putText(frame, "{}".format(status[RoI]),(box[RoI][0],box[RoI][1]), cv2.FONT_HERSHEY_TRIPLEX,.8, color, 2)   

                # Guide:
                cv2.putText(frame, "{}".format("S-5: P-Wall Set Station  "), (1000,20), cv2.FONT_HERSHEY_TRIPLEX,.4, (255, 0, 0), 1)   
                cv2.putText(frame, "{}".format("S-600: SideWall Set Station"), (1000,40), cv2.FONT_HERSHEY_TRIPLEX,.4, (255, 0, 0), 1)   
               
                # My information:
                cv2.putText(frame, "{}".format("This video is part of a Research by Roshan Panahi"), (10,640), cv2.FONT_HERSHEY_TRIPLEX,.4, (0, 255, 255), 1)   
                cv2.putText(frame, "{}".format("PanahiR@OregonState.Edu"), (10,660), cv2.FONT_HERSHEY_TRIPLEX,.4, (0, 255, 255), 1)   
                cv2.putText(frame, "{}".format("Adviser: Dr.Joseph Louis"), (10,680), cv2.FONT_HERSHEY_TRIPLEX,.4, (0, 255, 255), 1)   
               
                # CPM
                x,y=1150,80
                frame = cv2.rectangle(frame, (x,y), (x+50,y+30), (0, 0, 200), 2)  
                frame = cv2.rectangle(frame, (x-50,y+15), (x,y+16), (0, 0, 200), 2)  
                cv2.putText(frame, "{}".format("S-5"), (x,y+15), cv2.FONT_HERSHEY_TRIPLEX,.4,(0, 255, 0), 1)   

                frame = cv2.rectangle(frame, (x-100,y), (x-50,y+30), color, 2)    
                frame = cv2.rectangle(frame, (x-150,y+15), (x-100,y+16), color, 2)    
                cv2.putText(frame, "{}".format("S-600"), (x-100,y+15), cv2.FONT_HERSHEY_TRIPLEX,.4, (0, 255, 0), 1)   

                video_writer.write(frame)  

        counter+=1
    cap.release()
    cv2.destroyAllWindows()  
def show_video(image):
    cv2.imshow("image",image)
    cv2.waitKey(0)
def draw_match(image_retrieve,kp_retrieve,image_support,kp_support,match_index_support):
    Points_retrieve=[]
    Points_support=[]
    line_count=50
    buffer=50
    m0,n0,_=image_retrieve.shape
    m1,n1,_=image_support.shape
    
    for keypoint in kp_retrieve:
        Points_retrieve.append([int(keypoint.pt[0]),int(keypoint.pt[1])])
    for keypoint in kp_support:
        Points_support.append([int(keypoint.pt[0])+buffer+n0,int(keypoint.pt[1])])  
    
    
    canvas = np.zeros((m0, n0+buffer+n1+1, 3))
    canvas[0:m0,0:n0,:]=image_retrieve/255
    canvas[0:m1,n0+buffer+1:n0+buffer+n1+1,:]=image_support/255
    
    
    for i in range(np.minimum(len(Points_support),len(Points_retrieve))):
        cv2.circle(canvas, tuple(np.array(Points_retrieve[i])), radius=3, color=(0, 0, 255), thickness=-1)
        cv2.circle(canvas, tuple(np.array(Points_support[i])), radius=3, color=(0, 0, 255), thickness=-1)    

    for i in range(line_count):    
        cv2.line(canvas,tuple(np.array(Points_retrieve[i])),tuple(np.array(Points_support[match_index_support[i]])),(255,0,0),1)
    
    
    cv2.imshow('Match',canvas)
    cv2.waitKey()
def smooth(list, filter_size):
    list=scipy.signal.medfilt(list, kernel_size=filter_size)
    return(list)

def proccess(video):
    
    import torch
    import torch.nn as nn  # All neural network modules, nn.Linear, nn.Conv2d, BatchNorm, Loss functions
    import torch.optim as optim  # For all Optimization algorithms, SGD, Adam, etc.
    import torchvision.transforms as transforms  # Transformations we can perform on our dataset
    import torch.nn.functional as F
    import torchvision
    import os
    import pandas as pd
    from skimage import io
    from CustomDataset_Multiple_Box import BoxTrackingDataset
    from torch.utils.data import (
        Dataset,
        DataLoader,
    )  # Gives easier dataset managment and creates mini batches
    
    # Varibles
    global write_video_bool
    global write_excel_bool
    global draw_match_bool
    global draw_cost_bool
    global create_excel_bool
    global smoothing_bool
    global plot_target_bool
    global step_frame
    global cost
    global RoI_Count
    global Stations
    global batch_size

    Stations=[]
    keypoints_support=[]
    descriptors_support=[]
    support_images_all=[]
    target_all=[[],[]]
    target=[[],[]]
    cost=[[],[]]

    descriptors_support=[[],[]]
    
    draw_cost_bool=False
    draw_match_bool=False
    write_excel_bool=False
    write_video_bool=True
    create_excel_bool=False
    smoothing_bool=True
    plot_target_bool=True
    
    step_frame=1
    RoI_Count=2
    batch_size=12
    
    Stations.append('St-600')
    Stations.append('St-5')
    Stations.append('St-7')

    # Load Data
    dataset = BoxTrackingDataset(
        csv_file_Support="./Box_Annotation.csv",
        csv_file_RoI="./RoI.csv",
        video=video,
        support="./GT",
        transform=None,
        batch_size=batch_size,
        step_frame=step_frame
    )

    support_images_1, _,support_images_2, _ = dataset.get_support_images(raw = True)
    dataset.transform = None
    support_images_all.append(support_images_1)
    support_images_all.append(support_images_2)
    
    for RoI in range(RoI_Count): 
        hist0 = support_images_all[RoI][0]
        hist1 = support_images_all[RoI][1]
        tmp0, tmp1 = np.array([]), np.array([])
        for i in range(3):
            h0 = cv2.calcHist([hist0.numpy()], [i], None, [256], [0,256])
            h1 = cv2.calcHist([hist1.numpy()], [i], None, [256], [0,256])
            tmp0 = np.append(tmp0, h0) 
            tmp1 = np.append(tmp1, h1)                  
        descriptors_support[RoI].append(tmp0)      
        descriptors_support[RoI].append(tmp1)
    
    batch_number=0
    print(len(dataset))
    for i in range (0,len(dataset)):

        #batch_images_all=[[[],[]],[[],[]],[[],[]],[[],[]],[[],[]],[[],[]],[[],[]],[[],[]]]
        total_frame=len(dataset)
        batch_images_left,batch_images_right = dataset(i)
        n = len(batch_images_left)
        batch_images_all=[[],[]]*n
 
        for frame_in_batch in range(n):
           
            batch_images_all[0].append(batch_images_left[frame_in_batch])
            batch_images_all[1].append(batch_images_right[frame_in_batch])
            batch_number+=1
  
        if n > 0:
            compute_feature_matching(support_images_all,descriptors_support, batch_number, batch_images_all,total_frame,0,target,target_all)
        
        print('test')
        print(target)

    return(target)   
    
    
#================main=======================

if __name__ == "__main__":

    Input_path="G:/Surveillance_Videos_VBC/20201211/Segmented_videos/Camera_3_Day_1/"
    Input_path="D:/OtherCodes/BoxTracking/Videos/Surveillance/Inputs/test_6.mp4"
    Input_path= "G:/Surveillance_Videos_VBC/20201211/UnSegmented_videos/Camera_3_Day_5_DVR2_ch3_20201030055959_20201030170024.mp4"
    Input_path="D:/OtherCodes/BoxTracking/Videos/Surveillance/Inputs/test_3.mp4"
    Input_path = "./"
    p_1=proccess(Input_path)
    target_p_1=p_1

    if plot_target_bool:
        plt.yticks(np.arange(0, 1.1,1))
        plt.plot(target_p_1[0])
        plt.show() 
        plt.plot(target_p_1[1])
        plt.show() 
        
    if smoothing_bool:
        #for RoI in range(RoI_Count):
        target_p_1[0]=smooth(target_p_1[0], 15)
        target_p_1[1]=smooth(target_p_1[1], 15)

        plt.yticks(np.arange(0, 1.1,1))
        plt.plot(target_p_1[0])
        plt.plot(target_p_1[1])

    input_dir=Input_path
    output_dir='./out.avi'
    
    if write_video_bool:
        write_video(input_dir, output_dir, target_p_1)  
        
    if create_excel_bool:
        create_excel()
        
    if write_excel_bool:
       write_excel(target_p_1,50)
    